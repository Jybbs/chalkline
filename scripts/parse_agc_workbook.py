"""
Extract structured JSON from the AGC Maine stakeholder workbook.

Reads each sheet from `data/stakeholder/raw/agc-maine-2026.xlsx`, resolves
embedded hyperlinks to actual URLs, and writes one JSON file per logical
dataset into `data/stakeholder/reference/`.

    uv run python scripts/parse_agc_workbook.py
"""

import pandas as pd

from functools    import partial
from json         import dumps
from openpyxl     import load_workbook
from pathlib      import Path
from urllib.parse import parse_qs, urlparse


class StakeholderExtractor:
    """
    Extract and clean reference data from the AGC Maine Excel workbook.

    This is a scoping atlas, not the job posting corpus. It tells us which
    occupations, employers, apprenticeships, and educational programs are
    in scope. Each sheet becomes a JSON file in `reference/`.
    """

    _clean = staticmethod(lambda v: str(v).strip() if pd.notna(v) else "")

    def __init__(self, path: Path):
        """
        Load the workbook and configure output paths.

        Args:
            path: Path to the AGC Maine Excel workbook.
        """
        self._read_sheet = partial(pd.read_excel, pd.ExcelFile(path), header=None)
        self.output_dir  = path.parents[1] / "reference"
        self.wb          = load_workbook(path)

    # -----------------------------------------------------------------
    # Private Helpers
    # -----------------------------------------------------------------

    def _parse_job_board(self, section: pd.DataFrame) -> list[dict]:
        """
        Parse a job board section into a list of records.

        Args:
            section: DataFrame slice containing the job board rows.

        Returns:
            List of dicts with `category`, `name`, `focus`, and `best_for`
            keys.
        """
        return (
            section.dropna(subset=[1])
            .iloc[:, :4]
            .set_axis(["category", "name", "focus", "best_for"], axis=1)
            .map(self._clean)
            .to_dict(orient="records")
        )

    def _resolve_hyperlinks(self, col: int, sheet_name: str) -> dict[int, str]:
        """
        Map Excel row numbers to resolved hyperlink URLs for a column.

        `pandas` throws away embedded hyperlinks during `read_excel`, so
        we walk the `openpyxl` sheet directly to get them back. Google
        Search redirects are unwrapped along the way. Keys in the returned
        dict are 1-based Excel row numbers.

        Args:
            col        : 1-based column number to scan for hyperlinks.
            sheet_name : Excel sheet name to read from the workbook.

        Returns:
            Mapping from 1-based Excel row numbers to resolved URL strings.
        """
        return {
            cell.row: self._unwrap_google_redirect(h.target)
            for row in self.wb[sheet_name].iter_rows()
            if (h := (cell := row[col - 1]).hyperlink) and h.target
            and cell.row is not None
        }

    @staticmethod
    def _unwrap_google_redirect(url: str) -> str:
        """
        Return the real destination from a Google Search redirect URL.

        Parses the `q` parameter from `www.google.com/search?q=...`
        redirects. Non-Google URLs pass through unchanged.

        Args:
            url: Candidate URL that may be a Google redirect.

        Returns:
            The unwrapped destination URL, or the original if not a
            Google redirect.
        """
        if (
            (parsed := urlparse(url)).hostname == "www.google.com"
            and parsed.path == "/search"
            and (q := parse_qs(parsed.query).get("q"))
        ):
            return q[0]
        return url

    # -----------------------------------------------------------------
    # Extractors
    # -----------------------------------------------------------------

    def extract_agc_members(self) -> list[dict]:
        """
        Extract AGC Maine member companies with their type classification.

        The stakeholder flagged this list as private, so it should only be
        used for searching job postings and public data. About 23 companies
        also appear on the DOT prequal list.

        Returns:
            List of dicts with `type` and `name` keys.
        """
        return (
            self._read_sheet(
                names      = ["type", "name"],
                sheet_name = "AGC Member list ",
                skiprows   = 4
            )
            .dropna(subset=["name"])
            .map(self._clean)
            .to_dict(orient="records")
        )

    def extract_apprenticeships(self) -> list[dict]:
        """
        Extract AGC-sponsored registered apprenticeship programs.

        The RAPIDS codes here are DOL apprenticeship IDs, not O*NET codes.
        Term hours are strings because some are ranges like "4500 - 5000".
        Used downstream in CL-16 for pathway graphs.

        Returns:
            List of dicts with `rapids_code`, `title`, and `term_hours`
            keys.
        """
        return (
            self._read_sheet(
                names      = ["rapids_code", "title", "term_hours"],
                sheet_name = "AGC Sponsored Apprenticeships ",
                skiprows   = 2
            )
            .dropna(subset=["title"])
            .map(self._clean)
            .to_dict(orient="records")
        )

    def extract_cc_programs(self) -> dict:
        """
        Extract community college programs and workforce initiatives.

        The sheet has degree programs up top and workforce initiatives below
        with different columns, so this returns both under separate keys.
        Hyperlinks come from the Excel cell metadata.

        Returns:
            Dict with `degrees` and `initiatives` keys, each mapping to a
            list of record dicts.
        """
        sheet_name = "Community College Programs"
        urls       = self._resolve_hyperlinks(col=4, sheet_name=sheet_name)
        df         = self._read_sheet(sheet_name=sheet_name)

        def section(rows: pd.DataFrame, columns: list[str]) -> list[dict]:
            result = (
                rows.dropna(subset=[1])
                .iloc[:, :len(columns)]
                .set_axis(columns, axis=1)
                .map(self._clean)
            )
            result["url"] = (result.index + 1).map(urls).fillna("")
            return result.to_dict(orient="records")

        sections = {
            "degrees"     : (df.iloc[2:16], ["college", "program", "credential"]),
            "initiatives" : (df.iloc[20:],  ["initiative", "description", "best_for"])
        }
        return {
            name: section(rows, cols)
            for name, (rows, cols) in sections.items()
        }

    def extract_dot_contractors(self) -> list[dict]:
        """
        Extract MaineDOT prequalified contractors.

        Rows starting with an asterisk are footer notes and get dropped.
        Most website notes just repeat the company name rather than giving
        an actual URL.

        Returns:
            List of dicts with `company` and `website_note` keys.
        """
        return (
            self._read_sheet(
                names      = ["company", "website_note"],
                sheet_name = "JobBoards -DOT Prequal List",
                skiprows   = 1
            )
            .dropna(subset=["company"])
            .pipe(lambda df: df[~df["company"].str.startswith("*")])
            .map(self._clean)
            .to_dict(orient="records")
        )

    def extract_job_boards(self) -> dict:
        """
        Extract job board references split into Maine and national.

        Not especially useful on its own, more of a reference for where
        to look when we start collecting postings.

        Returns:
            Dict with `maine` and `national` keys, each mapping to a list
            of record dicts.
        """
        df = self._read_sheet(sheet_name="General Job Websites")
        return {
            "maine"    : self._parse_job_board(df.iloc[2:14]),
            "national" : self._parse_job_board(df.iloc[17:])
        }

    def extract_onet_codes(self) -> list[dict]:
        """
        Extract stakeholder-curated O*NET occupation codes.

        The 21 unique SOC codes here define the project's occupation scope
        and drive CL-04 lexicon curation. The sheet has three sector blocks
        with their own headers, and the original spreadsheet misspells
        "Heavy" as "Heay", which we fix.

        Returns:
            List of dicts with `soc_code`, `title`, `role_description`,
            and `sector` keys.
        """
        df = (
            self._read_sheet(sheet_name="Construction ONet Codes")
            .iloc[:, :3]
            .set_axis(["soc_code", "title", "role_description"], axis=1)
            .map(self._clean)
        )
        is_header = df["soc_code"].ne("") & df["title"].eq("")
        is_soc    = df["soc_code"].str.match(r"\d{2}-")

        df["sector"] = df.loc[is_header, "soc_code"].str.replace("Heay", "Heavy")
        df["sector"] = df["sector"].ffill()

        return (
            df.loc[is_soc]
            .drop_duplicates(subset=["soc_code"])
            .to_dict(orient="records")
        )

    def extract_umaine_programs(self) -> list[dict]:
        """
        Extract University of Maine System construction and engineering
        programs.

        Four campuses, ranging from 1+3 articulation pathways up to full
        B.S. and B.Arch degrees. Hyperlinks come from the Excel cell
        metadata, same as `extract_cc_programs`.

        Returns:
            List of dicts with `campus`, `category`, `degree`, `program`,
            and `url` keys.
        """
        sheet_name = "Umaine Programs"
        urls       = self._resolve_hyperlinks(col=5, sheet_name=sheet_name)
        df = (
            self._read_sheet(
                names      = ["campus", "category", "degree", "program"],
                sheet_name = sheet_name,
                skiprows   = 3,
                usecols    = range(4)
            )
            .dropna(subset=["program"])
            .map(self._clean)
        )
        df["url"] = (df.index + 4).map(urls).fillna("")
        return df.to_dict(orient="records")

    # -----------------------------------------------------------------
    # Runner
    # -----------------------------------------------------------------

    def run_all(self):
        """
        Extract all sheets and write JSON files to the reference directory.
        """
        self.output_dir.mkdir(exist_ok=True)

        for method in (
            self.extract_agc_members,
            self.extract_apprenticeships,
            self.extract_cc_programs,
            self.extract_dot_contractors,
            self.extract_job_boards,
            self.extract_onet_codes,
            self.extract_umaine_programs
        ):
            path = self.output_dir / f"{method.__name__.removeprefix('extract_')}.json"
            path.write_text(dumps(method(), ensure_ascii=False, indent=2) + "\n")


if __name__ == "__main__":

    StakeholderExtractor(
        path=(
            Path(__file__).resolve().parents[1]
            / "data/stakeholder/raw/agc-maine-2026.xlsx"
        )
    ).run_all()
